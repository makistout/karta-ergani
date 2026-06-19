"""Εξαγωγή & parsing Excel από ASP.NET grid portal Ergani (μόνο .xlsx/.xls)."""

from __future__ import annotations

import io
import re
from urllib.parse import urljoin

import requests

from app.portal_schedule_sync import REQUEST_TIMEOUT, _extract_aspnet_form_data

EXCEL_EXPORT_TIMEOUT = 300
EXCEL_EXPORT_ARGUMENT = "ExcelExport$1"


def _is_excel_content(content: bytes, content_type: str) -> bool:
    ctype = (content_type or "").lower()
    head = content[:8]
    if head.startswith(b"PK"):
        return True
    if head.startswith(b"\xd0\xcf\x11\xe0"):
        return True
    return "spreadsheetml" in ctype or "openxmlformats" in ctype or "ms-excel" in ctype


def download_grid_excel(
    session: requests.Session,
    html: str,
    page_url: str,
    *,
    grid_event_target: str,
    excel_argument: str = EXCEL_EXPORT_ARGUMENT,
) -> tuple[bytes, str]:
    """POST __doPostBack ExcelExport — επιστρέφει (bytes, content-type)."""
    data = _extract_aspnet_form_data(html, include_text=True)
    data["__EVENTTARGET"] = grid_event_target
    data["__EVENTARGUMENT"] = excel_argument
    for key in list(data.keys()):
        if key.endswith("$SearchControlSearchButton"):
            del data[key]

    action = page_url
    m = re.search(r'<form[^>]+action="([^"]*)"', html, re.I)
    if m:
        action = urljoin(page_url, m.group(1))

    resp = session.post(
        action,
        data=data,
        timeout=EXCEL_EXPORT_TIMEOUT,
        allow_redirects=True,
    )
    if resp.status_code >= 400:
        raise RuntimeError(f"Αποτυχία Excel export — HTTP {resp.status_code}")
    if "error.aspx" in resp.url.lower():
        raise RuntimeError("Αποτυχία Excel export — error.aspx")
    ctype = (resp.headers.get("Content-Type") or "").split(";")[0].strip().lower()
    body = resp.content
    if not body:
        raise RuntimeError("Κενό αρχείο Excel export")
    if not _is_excel_content(body, ctype):
        raise RuntimeError("Το Excel export δεν επέστρεψε αρχείο Excel (.xlsx/.xls)")
    return body, ctype


def _parse_xlsx(content: bytes) -> list[list[str]]:
    from openpyxl import load_workbook

    # read_only=False — το export Ergani έχει λανθασμένο dimension ref="A1"
    wb = load_workbook(io.BytesIO(content), read_only=False, data_only=True)
    try:
        ws = wb.active
        rows: list[list[str]] = []
        for row in ws.iter_rows(values_only=True):
            cells = ["" if v is None else str(v).strip() for v in row]
            if any(cells):
                rows.append(cells)
        return rows
    finally:
        wb.close()


def _parse_xls(content: bytes) -> list[list[str]]:
    import xlrd

    book = xlrd.open_workbook(file_contents=content)
    sheet = book.sheet_by_index(0)
    rows: list[list[str]] = []
    for r in range(sheet.nrows):
        cells = [str(sheet.cell_value(r, c)).strip() for c in range(sheet.ncols)]
        if any(cells):
            rows.append(cells)
    return rows


def _parse_excel_bytes(content: bytes, content_type: str = "") -> list[list[str]]:
    """Ανάγνωση μόνο πραγματικού Excel — όχι HTML."""
    ctype = (content_type or "").lower()
    head = content[:8]
    if head.startswith(b"PK") or "spreadsheetml" in ctype or "openxmlformats" in ctype:
        return _parse_xlsx(content)
    if head.startswith(b"\xd0\xcf\x11\xe0") or "ms-excel" in ctype:
        return _parse_xls(content)
    raise RuntimeError("Μη αναγνωρισμένο αρχείο Excel")


def _header_index(headers: list[str], *needles: str) -> int | None:
    for i, h in enumerate(headers):
        norm = (h or "").upper().replace(" ", "")
        for n in needles:
            if n in norm:
                return i
    return None


def _work_log_row_from_cells(
    cells: list[str],
    col: dict[str, int],
    *,
    default_branch_aa: str = "",
) -> list[str] | None:
    def pick(key: str, default: int) -> str:
        idx = col.get(key, default)
        if idx < 0:
            return ""
        return (cells[idx] if idx < len(cells) else "").strip()

    if "aa" in col:
        aa = pick("aa", 0)
    else:
        aa = str(default_branch_aa or "").strip()
    afm = re.sub(r"\D", "", pick("afm", 1))
    if not re.fullmatch(r"\d{8,11}", afm):
        return None
    eponymo = pick("eponymo", 2)
    onoma = pick("onoma", 3)
    work_date = pick("date", 4)
    hour_from = pick("from", 5)
    hour_to = pick("to", 6)
    return [aa, afm, eponymo, onoma, work_date, hour_from, hour_to]


def _detect_work_log_columns(header: list[str]) -> dict[str, int]:
    col: dict[str, int] = {}
    mapping = {
        "aa": ("ΑΑΠΑΡΑΡΤΗΜΑΤΟΣ",),
        "afm": ("ΑΦΜ", "AFM"),
        "eponymo": ("ΕΠΩΝΥΜΟ", "EPONIMO"),
        "onoma": ("ΟΝΟΜΑ", "ONOMA"),
        "date": ("ΗΜ/ΝΙΑ", "ΗΜΕΡΟΜΗΝΙΑ", "DATE"),
        "from": ("ΩΡΑΑΠΟ", "ΩΡΑΑΠΌ", "HOURFROM", "ΑΠΟ", "ΑΠΌ"),
        "to": ("ΩΡΑΕΩΣ", "HOURTO", "ΕΩΣ", "ΈΩΣ"),
    }
    for key, needles in mapping.items():
        idx = _header_index(header, *needles)
        if idx is not None:
            col[key] = idx
    return col


def normalize_work_log_grid_rows(
    raw_rows: list[list[str]],
    *,
    default_branch_aa: str = "",
) -> list[list[str]]:
    """Μετατροπή γραμμών Excel → μορφή 7 κελιών (ίδια με HTML grid fallback)."""
    if not raw_rows:
        return []

    header_row: list[str] | None = None
    col: dict[str, int] = {}
    start = 0
    for i, row in enumerate(raw_rows[:5]):
        if _header_index(row, "ΑΦΜ", "AFM") is not None:
            header_row = row
            col = _detect_work_log_columns(row)
            start = i + 1
            break

    if not col:
        col = {"afm": 0, "eponymo": 1, "onoma": 2, "date": 3, "from": 4, "to": 5}

    out: list[list[str]] = []
    for row in raw_rows[start:]:
        if not row or all(not (c or "").strip() for c in row):
            continue
        if header_row and row == header_row:
            continue
        norm = _work_log_row_from_cells(
            row, col, default_branch_aa=default_branch_aa
        )
        if norm:
            out.append(norm)
    return out


def parse_work_log_export(
    content: bytes,
    content_type: str = "",
    *,
    default_branch_aa: str = "",
) -> list[list[str]]:
    """Parse μόνο Excel export πραγματικής απασχόλησης → grid rows."""
    raw = _parse_excel_bytes(content, content_type)
    return normalize_work_log_grid_rows(raw, default_branch_aa=default_branch_aa)


def fetch_work_log_rows_via_excel(
    session: requests.Session,
    html: str,
    page_url: str,
    *,
    grid_event_target: str,
    default_branch_aa: str = "",
) -> list[list[str]]:
    content, ctype = download_grid_excel(
        session, html, page_url, grid_event_target=grid_event_target
    )
    rows = parse_work_log_export(
        content, ctype, default_branch_aa=default_branch_aa
    )
    if not rows:
        raise RuntimeError("Το Excel export δεν περιέχει εγγραφές πραγματικής")
    return rows


def _schedule_row_from_cells(
    cells: list[str],
    col: dict[str, int],
    *,
    default_branch_aa: str = "",
) -> list[str] | None:
    def pick(key: str, default: int) -> str:
        idx = col.get(key, default)
        if idx < 0:
            return ""
        return (cells[idx] if idx < len(cells) else "").strip()

    if "aa" in col:
        aa = pick("aa", 0)
    else:
        aa = str(default_branch_aa or "").strip()
    afm = re.sub(r"\D", "", pick("afm", 1))
    if not re.fullmatch(r"\d{8,11}", afm):
        return None
    onoma = pick("onoma", 2)
    eponymo = pick("eponymo", 3)
    work_date = pick("date", 4)
    po = pick("po", 5)
    card = pick("card", 6)
    break_txt = pick("break", 7)
    employment = pick("employment", 8)
    return [aa, afm, onoma, eponymo, work_date, po, card, break_txt, employment]


def _detect_schedule_columns(header: list[str]) -> dict[str, int]:
    col: dict[str, int] = {}
    mapping = {
        "aa": ("ΑΑΠΑΡΑΡΤΗΜΑΤΟΣ",),
        "afm": ("ΑΦΜ", "AFM"),
        "onoma": ("ΟΝΟΜΑ", "ONOMA"),
        "eponymo": ("ΕΠΩΝΥΜΟ", "EPONIMO"),
        "date": ("ΗΜ/ΝΙΑ", "ΗΜΕΡΟΜΗΝΙΑ", "DATE"),
        "po": ("ΨΗΦΙΑΚΗΟΡΓΑΝΩΣΗ", "ΨΗΦΙΑΚΗ", "ΨΟ"),
        "card": ("ΚΑΡΤΑΕΡΓΑΣΙΑΣ", "ΚΑΡΤΑ"),
        "break": ("ΔΙΑΛΕΙΜΜΑ", "BREAK"),
        "employment": ("ΑΠΑΣΧΟΛΗΣΗ", "EMPLOYMENT"),
    }
    for key, needles in mapping.items():
        idx = _header_index(header, *needles)
        if idx is not None:
            col[key] = idx
    return col


def normalize_schedule_grid_rows(
    raw_rows: list[list[str]],
    *,
    default_branch_aa: str = "",
) -> list[list[str]]:
    """Μετατροπή γραμμών Excel → μορφή 9 κελιών (ίδια με HTML grid fallback)."""
    if not raw_rows:
        return []

    header_row: list[str] | None = None
    col: dict[str, int] = {}
    start = 0
    for i, row in enumerate(raw_rows[:5]):
        if _header_index(row, "ΑΦΜ", "AFM") is not None:
            header_row = row
            col = _detect_schedule_columns(row)
            start = i + 1
            break

    if not col:
        col = {
            "afm": 0,
            "onoma": 1,
            "eponymo": 2,
            "date": 3,
            "po": 4,
            "card": 5,
            "break": 6,
            "employment": 7,
        }

    out: list[list[str]] = []
    for row in raw_rows[start:]:
        if not row or all(not (c or "").strip() for c in row):
            continue
        if header_row and row == header_row:
            continue
        norm = _schedule_row_from_cells(
            row, col, default_branch_aa=default_branch_aa
        )
        if norm:
            out.append(norm)
    return out


def parse_schedule_export(
    content: bytes,
    content_type: str = "",
    *,
    default_branch_aa: str = "",
) -> list[list[str]]:
    """Parse μόνο Excel export ψηφιακού ωραρίου → grid rows."""
    raw = _parse_excel_bytes(content, content_type)
    return normalize_schedule_grid_rows(raw, default_branch_aa=default_branch_aa)


def fetch_schedule_rows_via_excel(
    session: requests.Session,
    html: str,
    page_url: str,
    *,
    grid_event_target: str,
    default_branch_aa: str = "",
) -> list[list[str]]:
    content, ctype = download_grid_excel(
        session, html, page_url, grid_event_target=grid_event_target
    )
    rows = parse_schedule_export(content, ctype, default_branch_aa=default_branch_aa)
    if not rows:
        raise RuntimeError("Το Excel export δεν περιέχει εγγραφές ωραρίου")
    return rows
