"""Ανάγνωση Excel template εβδομαδιαίου ωραρίου και σύγκριση με τρέχον ψηφ. ωράριο."""

from __future__ import annotations

import json
import re
from datetime import date, datetime, time, timedelta
from typing import Any

from openpyxl import load_workbook

from app.repo_entities import list_employees_for_employer
from app.repo_schedule import list_schedule_for_store
from app.schedule_excel_layout import (
    BASE_HEADERS,
    DAY_FIELD_COUNT,
    INSTRUCTIONS_SHEET,
    LEGACY_DAY_HEADERS,
    LEGACY_WIDE_DAY_FIELD_COUNT,
    ROWS_PER_EMPLOYEE,
    SINGLE_SHEET_DATA_START_ROW,
    SINGLE_SHEET_HEADER_ROW_FIELDS,
    WEEK_SHEET,
    single_sheet_day_col,
)
from app.work_card_payload import norm_afm

HEADERS = LEGACY_DAY_HEADERS
_DATE_IN_TITLE = re.compile(r"(\d{1,2}/\d{1,2}/\d{4})")
_WEEK_RANGE = re.compile(
    r"(\d{1,2}/\d{1,2}/\d{4})\s*-\s*(\d{1,2}/\d{1,2}/\d{4})"
)


def _hm_short(value: str | None) -> str:
    m = re.match(r"^(\d{1,2}):(\d{2})", str(value or "").strip())
    if not m:
        return ""
    return f"{int(m.group(1)):02d}:{m.group(2)}"


def _hm_from_digits(raw: str) -> str:
    """Μετατροπή 1–4 ψηφίων σε ΩΩ:ΛΛ (π.χ. 900→09:00, 1340→13:40)."""
    digits = re.sub(r"\D", "", str(raw or ""))
    if not digits or len(digits) > 4:
        return ""
    digits = digits.zfill(4)
    h, m = int(digits[:2]), int(digits[2:])
    if h > 23 or m > 59:
        return ""
    return f"{h:02d}:{m:02d}"


def _format_time_cell(value: Any) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, time):
        return f"{value.hour:02d}:{value.minute:02d}"
    if isinstance(value, datetime):
        return f"{value.hour:02d}:{value.minute:02d}"
    if isinstance(value, (int, float)):
        f = float(value)
        # Κλάσμα ημέρας Excel (π.χ. 0.375 = 09:00) από παλιά templates hh:mm.
        if 0 <= f < 1:
            total = int(round(f * 24 * 60))
            total %= 24 * 60
            h, m = divmod(total, 60)
            return f"{h:02d}:{m:02d}"
        # Ακέραια ψηφία HHMM με format 00\:00 (π.χ. 900, 1340).
        if f == int(f):
            return _hm_from_digits(str(int(f)))
        return ""
    text = str(value).strip()
    as_hm = _hm_short(text)
    if as_hm:
        return as_hm
    return _hm_from_digits(text)


def _parse_ergani_date(text: str | None) -> str | None:
    s = str(text or "").strip()
    m = _DATE_IN_TITLE.search(s)
    if not m:
        return None
    try:
        dt = datetime.strptime(m.group(1), "%d/%m/%Y").date()
        return dt.strftime("%d/%m/%Y")
    except ValueError:
        return None


def _schedule_is_rest(row: dict[str, Any]) -> bool:
    shift = str(row.get("shift_type") or "").strip().upper()
    if shift in ("ΑΝ", "AN", "Ρ", "ΡΕΠΟ"):
        return True
    if re.search(r"ρεπο|ανάπαυση", shift, re.I):
        return True
    hf = _hm_short(str(row.get("hour_from") or ""))
    ht = _hm_short(str(row.get("hour_to") or ""))
    return not hf and not ht and bool(shift)


def _intervals_from_schedule_rows(rows: list[dict[str, Any]]) -> list[dict[str, str]]:
    out: list[dict[str, str]] = []
    for row in rows:
        if _schedule_is_rest(row):
            return []
        hf = _hm_short(str(row.get("hour_from") or ""))
        ht = _hm_short(str(row.get("hour_to") or ""))
        if hf or ht:
            out.append({"hour_from": hf, "hour_to": ht})
    return out


def _snapshot_from_intervals(
    *,
    intervals: list[dict[str, str]],
    is_rest: bool,
    schedule_type: str = "ΕΡΓ",
) -> list[dict[str, Any]]:
    if is_rest:
        return [
            {
                "hour_from": None,
                "hour_to": None,
                "shift_type": "ΑΝΑΠΑΥΣΗ/ΡΕΠΟ",
                "schedule_type": "ΑΝ",
            }
        ]
    return [
        {
            "hour_from": item.get("hour_from") or None,
            "hour_to": item.get("hour_to") or None,
            "shift_type": schedule_type,
            "schedule_type": schedule_type,
        }
        for item in intervals
    ]


def _snapshots_equal(current: list[dict[str, Any]], proposed: list[dict[str, Any]]) -> bool:
    def norm(block: list[dict[str, Any]]) -> list[tuple[str | None, str | None, str | None]]:
        rows: list[tuple[str | None, str | None, str | None]] = []
        for item in block:
            rows.append(
                (
                    _hm_short(str(item.get("hour_from") or "")) or None,
                    _hm_short(str(item.get("hour_to") or "")) or None,
                    str(item.get("schedule_type") or item.get("shift_type") or "").strip().upper() or None,
                )
            )
        return sorted(rows)

    return norm(current) == norm(proposed)


def _current_snapshot_for_employee(
    schedule_rows: list[dict[str, Any]],
    employee_afm: str,
) -> list[dict[str, Any]]:
    emp = norm_afm(employee_afm)
    emp_rows = [row for row in schedule_rows if norm_afm(row.get("employee_afm") or "") == emp]
    if not emp_rows:
        return []
    if any(_schedule_is_rest(row) for row in emp_rows):
        return _snapshot_from_intervals(intervals=[], is_rest=True)
    intervals = _intervals_from_schedule_rows(emp_rows)
    return _snapshot_from_intervals(intervals=intervals, is_rest=False)


def _build_import_row(
    *,
    row_no: int,
    sheet_name: str,
    work_date: str,
    afm: str,
    eponymo: str | None,
    onoma: str | None,
    import_action: str,
    intervals: list[dict[str, str]],
    schedule_type: str | None,
    current_schedule: list[dict[str, Any]],
    row_errors: list[str] | None = None,
) -> dict[str, Any]:
    errors = list(row_errors or [])
    current_snapshot = _current_snapshot_for_employee(current_schedule, afm)
    if import_action == "skip":
        proposed_snapshot = current_snapshot
        change_kind = "skip"
    elif import_action in ("rest", "absent"):
        proposed_snapshot = _snapshot_from_intervals(intervals=[], is_rest=True)
        if errors:
            change_kind = "error"
        elif not current_snapshot:
            change_kind = "new"
        elif _snapshots_equal(current_snapshot, proposed_snapshot):
            change_kind = "same"
        else:
            change_kind = "update"
    else:
        proposed_snapshot = _snapshot_from_intervals(
            intervals=intervals,
            is_rest=False,
            schedule_type=schedule_type or "ΕΡΓ",
        )
        if errors:
            change_kind = "error"
        elif not current_snapshot:
            change_kind = "new"
        elif _snapshots_equal(current_snapshot, proposed_snapshot):
            change_kind = "same"
        else:
            change_kind = "update"

    hf1 = intervals[0].get("hour_from") if len(intervals) > 0 else None
    ht1 = intervals[0].get("hour_to") if len(intervals) > 0 else None
    hf2 = intervals[1].get("hour_from") if len(intervals) > 1 else None
    ht2 = intervals[1].get("hour_to") if len(intervals) > 1 else None
    return {
        "row_no": row_no,
        "sheet_name": sheet_name,
        "work_date": work_date,
        "employee_afm": afm,
        "eponymo": eponymo,
        "onoma": onoma,
        "import_action": import_action,
        "hour_from_1": hf1 or None,
        "hour_to_1": ht1 or None,
        "hour_from_2": hf2 or None,
        "hour_to_2": ht2 or None,
        "schedule_type": schedule_type,
        "change_kind": change_kind,
        "current_snapshot": current_snapshot,
        "proposed_snapshot": proposed_snapshot,
        "validation_errors": errors,
    }


def _employee_name_from_schedule(
    schedule_rows: list[dict[str, Any]],
    employee_afm: str,
) -> tuple[str | None, str | None]:
    emp = norm_afm(employee_afm)
    for row in schedule_rows:
        if norm_afm(row.get("employee_afm") or "") != emp:
            continue
        eponymo = str(row.get("eponymo") or "").strip() or None
        onoma = str(row.get("onoma") or "").strip() or None
        if eponymo or onoma:
            return eponymo, onoma
    return None, None


def _afms_missing_from_sheet(
    *,
    sheet_afms: set[str],
    known_afms: set[str],
    current_schedule: list[dict[str, Any]],
) -> set[str]:
    scheduled_afms = {
        norm_afm(row.get("employee_afm") or "")
        for row in current_schedule
        if norm_afm(row.get("employee_afm") or "")
    }
    universe = known_afms | scheduled_afms
    return {afm for afm in universe if afm not in sheet_afms}


def _validate_proposed_row(
    *,
    import_action: str,
    intervals: list[dict[str, str]],
) -> list[str]:
    errors: list[str] = []
    if import_action == "skip":
        return errors
    if import_action in ("rest", "absent"):
        return errors
    if not intervals:
        errors.append("Λείπουν ώρες εργασίας")
        return errors
    for idx, item in enumerate(intervals, start=1):
        hf = item.get("hour_from") or ""
        ht = item.get("hour_to") or ""
        if not hf or not ht:
            errors.append(f"Ατελές διάστημα {idx}: απαιτούνται Από και Έως")
    return errors


def _build_intervals(
    hour_from_1: str,
    hour_to_1: str,
    hour_from_2: str,
    hour_to_2: str,
) -> list[dict[str, str]]:
    intervals: list[dict[str, str]] = []
    if hour_from_1 or hour_to_1:
        intervals.append({"hour_from": hour_from_1, "hour_to": hour_to_1})
    if hour_from_2 or hour_to_2:
        intervals.append({"hour_from": hour_from_2, "hour_to": hour_to_2})
    return intervals


def _read_instructions(ws: Any) -> dict[str, Any]:
    instructions_week_label = str(ws["A2"].value or "").strip() or None
    week_from = week_to = None
    m = _WEEK_RANGE.search(instructions_week_label or "")
    if m:
        week_from, week_to = m.group(1), m.group(2)
    store_id_raw = ws["B16"].value
    try:
        store_id = int(store_id_raw) if store_id_raw is not None else None
    except (TypeError, ValueError):
        store_id = None
    return {
        "week_label": instructions_week_label,
        "instructions_week_label": instructions_week_label,
        "week_from": week_from,
        "week_to": week_to,
        "store_id": store_id,
        "employer_afm": str(ws["B17"].value or "").strip() or None,
        "branch_aa": str(ws["B18"].value or "").strip() or None,
    }


def _week_label_from_dates(dates: list[str]) -> tuple[str | None, str | None, str | None]:
    parsed: list[date] = []
    for value in dates:
        try:
            parsed.append(datetime.strptime(str(value).strip(), "%d/%m/%Y").date())
        except ValueError:
            continue
    if not parsed:
        return None, None, None
    start = min(parsed)
    end = max(parsed)
    week_from = start.strftime("%d/%m/%Y")
    week_to = end.strftime("%d/%m/%Y")
    return f"Εβδομάδα {week_from} - {week_to}", week_from, week_to


def _monday_from_meta(meta: dict[str, Any]) -> date | None:
    week_from = meta.get("week_from")
    if week_from:
        try:
            return datetime.strptime(str(week_from).strip(), "%d/%m/%Y").date()
        except ValueError:
            pass
    label = str(meta.get("week_label") or meta.get("instructions_week_label") or "").strip()
    m = _WEEK_RANGE.search(label)
    if m:
        try:
            return datetime.strptime(m.group(1), "%d/%m/%Y").date()
        except ValueError:
            pass
    return None


def _parse_excel_day_entry(
    *,
    action_raw: str,
    hf1: str,
    ht1: str,
    hf2: str,
    ht2: str,
    afm: str,
    known_afms: set[str],
) -> tuple[str, list[dict[str, str]], str | None, list[str]]:
    row_errors: list[str] = []
    if afm not in known_afms:
        row_errors.append(f"Άγνωστο ΑΦΜ {afm} για το κατάστημα")

    action = str(action_raw or "").strip().upper()
    if action == "ΡΕΠΟ":
        if any((hf1, ht1, hf2, ht2)):
            row_errors.append("Σε ΡΕΠΟ δεν πρέπει να υπάρχουν ώρες")
        return "rest", [], "ΑΝ", row_errors
    if any((hf1, ht1, hf2, ht2)):
        intervals = _build_intervals(hf1, ht1, hf2, ht2)
        row_errors.extend(
            _validate_proposed_row(import_action="work", intervals=intervals)
        )
        return "work", intervals, "ΕΡΓ", row_errors
    return "skip", [], None, row_errors


def _is_legacy_wide_single_sheet(ws: Any) -> bool:
    """Παλιό format: Από1/Έως1/Από2/Έως2 στην ίδια γραμμή."""
    label = str(
        ws.cell(SINGLE_SHEET_HEADER_ROW_FIELDS, single_sheet_day_col(0, 1, fields_per_day=3)).value
        or ""
    ).strip().upper().replace(" ", "")
    # Στο wide format η στήλη 5 είναι «Από1»· στο stacked είναι «Από».
    # Ελέγχουμε και τη στήλη του Από2 στο wide (col 7 με 5 fields).
    apo2 = str(
        ws.cell(
            SINGLE_SHEET_HEADER_ROW_FIELDS,
            single_sheet_day_col(0, 3, fields_per_day=LEGACY_WIDE_DAY_FIELD_COUNT),
        ).value
        or ""
    ).strip().upper().replace(" ", "")
    return label.startswith("ΑΠΟ1") or apo2.startswith("ΑΠΟ2")


def _parse_single_week_sheet(
    wb: Any,
    *,
    meta: dict[str, Any],
    employer_afm: str,
    branch_aa: str,
    employees: dict[str, dict[str, Any]],
    known_afms: set[str],
) -> dict[str, Any]:
    monday = _monday_from_meta(meta)
    if not monday:
        raise ValueError(
            "Το φύλλο «Οδηγίες» δεν έχει έγκυρη εβδομάδα (A2) — απαιτείται ημερομηνία Δευτέρας"
        )

    ws = wb[WEEK_SHEET]
    errors: list[str] = []
    work_dates = [(monday + timedelta(days=i)).strftime("%d/%m/%Y") for i in range(7)]
    schedule_by_date = {
        wd: list_schedule_for_store(employer_afm, branch_aa, wd) for wd in work_dates
    }

    base_headers = [str(ws.cell(SINGLE_SHEET_HEADER_ROW_FIELDS, c).value or "").strip() for c in range(1, 4)]
    if base_headers != BASE_HEADERS:
        errors.append(
            f"Το φύλλο «{WEEK_SHEET}» δεν έχει τα αναμενόμενα headers "
            f"({', '.join(BASE_HEADERS)}) στη γραμμή {SINGLE_SHEET_HEADER_ROW_FIELDS}"
        )

    legacy_wide = _is_legacy_wide_single_sheet(ws)
    fields_per_day = LEGACY_WIDE_DAY_FIELD_COUNT if legacy_wide else DAY_FIELD_COUNT
    parsed_rows: list[dict[str, Any]] = []
    row_no = 0
    max_row = ws.max_row or SINGLE_SHEET_DATA_START_ROW
    r = SINGLE_SHEET_DATA_START_ROW

    while r <= max_row:
        afm_raw = str(ws.cell(r, 1).value or "").strip()
        if not afm_raw:
            r += 1
            continue
        afm = norm_afm(afm_raw)
        eponymo = str(ws.cell(r, 2).value or "").strip()
        onoma = str(ws.cell(r, 3).value or "").strip()
        r2 = r + 1 if not legacy_wide else r

        for day_idx, work_date in enumerate(work_dates):
            energia_col = single_sheet_day_col(day_idx, 0, fields_per_day=fields_per_day)
            action_raw = str(ws.cell(r, energia_col).value or "")
            if legacy_wide:
                hf1 = _format_time_cell(
                    ws.cell(r, single_sheet_day_col(day_idx, 1, fields_per_day=fields_per_day)).value
                )
                ht1 = _format_time_cell(
                    ws.cell(r, single_sheet_day_col(day_idx, 2, fields_per_day=fields_per_day)).value
                )
                hf2 = _format_time_cell(
                    ws.cell(r, single_sheet_day_col(day_idx, 3, fields_per_day=fields_per_day)).value
                )
                ht2 = _format_time_cell(
                    ws.cell(r, single_sheet_day_col(day_idx, 4, fields_per_day=fields_per_day)).value
                )
            else:
                apo_col = single_sheet_day_col(day_idx, 1, fields_per_day=fields_per_day)
                eos_col = single_sheet_day_col(day_idx, 2, fields_per_day=fields_per_day)
                hf1 = _format_time_cell(ws.cell(r, apo_col).value)
                ht1 = _format_time_cell(ws.cell(r, eos_col).value)
                hf2 = _format_time_cell(ws.cell(r2, apo_col).value) if r2 <= max_row else ""
                ht2 = _format_time_cell(ws.cell(r2, eos_col).value) if r2 <= max_row else ""

            import_action, intervals, schedule_type, row_errors = _parse_excel_day_entry(
                action_raw=action_raw,
                hf1=hf1,
                ht1=ht1,
                hf2=hf2,
                ht2=ht2,
                afm=afm,
                known_afms=known_afms,
            )
            row_no += 1
            parsed_rows.append(
                _build_import_row(
                    row_no=row_no,
                    sheet_name=WEEK_SHEET,
                    work_date=work_date,
                    afm=afm,
                    eponymo=eponymo or (employees.get(afm) or {}).get("eponymo"),
                    onoma=onoma or (employees.get(afm) or {}).get("onoma"),
                    import_action=import_action,
                    intervals=intervals,
                    schedule_type=schedule_type,
                    current_schedule=schedule_by_date[work_date],
                    row_errors=row_errors,
                )
            )

        r += 1 if legacy_wide else ROWS_PER_EMPLOYEE

    actual_label, actual_from, actual_to = _week_label_from_dates(work_dates)
    if actual_label:
        instructions_label = str(meta.get("instructions_week_label") or meta.get("week_label") or "").strip()
        meta["week_label"] = actual_label
        meta["week_from"] = actual_from
        meta["week_to"] = actual_to
        if instructions_label and instructions_label != actual_label:
            errors.append(
                "Η εβδομάδα στο φύλλο «Οδηγίες» "
                f"({instructions_label}) διαφέρει από την αναμενόμενη "
                f"({actual_label}) — χρησιμοποιούνται οι ημερομηνίες του template."
            )

    return {
        "meta": meta,
        "work_dates": work_dates,
        "rows": parsed_rows,
        "errors": errors,
        "summary": summarize_import_rows(parsed_rows),
    }


def _parse_legacy_multi_sheet(
    wb: Any,
    *,
    meta: dict[str, Any],
    employer_afm: str,
    branch_aa: str,
    employees: dict[str, dict[str, Any]],
    known_afms: set[str],
) -> dict[str, Any]:
    parsed_rows: list[dict[str, Any]] = []
    errors: list[str] = []
    work_dates: list[str] = []
    row_no = 0

    for sheet_name in wb.sheetnames:
        if sheet_name == INSTRUCTIONS_SHEET:
            continue
        ws = wb[sheet_name]
        work_date = _parse_ergani_date(str(ws["A1"].value or ""))
        if not work_date:
            errors.append(f"Το φύλλο «{sheet_name}» δεν έχει έγκυρη ημερομηνία στον τίτλο (A1)")
            continue
        work_dates.append(work_date)
        current_schedule = list_schedule_for_store(employer_afm, branch_aa, work_date)

        header_vals = [str(ws.cell(3, c).value or "").strip() for c in range(1, 9)]
        if header_vals[:4] != HEADERS[:4]:
            errors.append(f"Το φύλλο «{sheet_name}» δεν έχει τα αναμενόμενα headers στη γραμμή 3")

        sheet_afms: set[str] = set()

        for r in range(4, ws.max_row + 1):
            afm_raw = str(ws.cell(r, 1).value or "").strip()
            if not afm_raw:
                continue
            row_no += 1
            afm = norm_afm(afm_raw)
            sheet_afms.add(afm)
            eponymo = str(ws.cell(r, 2).value or "").strip()
            onoma = str(ws.cell(r, 3).value or "").strip()
            action = str(ws.cell(r, 4).value or "").strip().upper()
            hf1 = _format_time_cell(ws.cell(r, 5).value)
            ht1 = _format_time_cell(ws.cell(r, 6).value)
            hf2 = _format_time_cell(ws.cell(r, 7).value)
            ht2 = _format_time_cell(ws.cell(r, 8).value)

            import_action, intervals, schedule_type, row_errors = _parse_excel_day_entry(
                action_raw=action,
                hf1=hf1,
                ht1=ht1,
                hf2=hf2,
                ht2=ht2,
                afm=afm,
                known_afms=known_afms,
            )

            parsed_rows.append(
                _build_import_row(
                    row_no=row_no,
                    sheet_name=sheet_name,
                    work_date=work_date,
                    afm=afm,
                    eponymo=eponymo or (employees.get(afm) or {}).get("eponymo"),
                    onoma=onoma or (employees.get(afm) or {}).get("onoma"),
                    import_action=import_action,
                    intervals=intervals,
                    schedule_type=schedule_type,
                    current_schedule=current_schedule,
                    row_errors=row_errors,
                )
            )

        for afm in sorted(
            _afms_missing_from_sheet(
                sheet_afms=sheet_afms,
                known_afms=known_afms,
                current_schedule=current_schedule,
            )
        ):
            row_no += 1
            emp = employees.get(afm) or {}
            sched_ep, sched_on = _employee_name_from_schedule(current_schedule, afm)
            parsed_rows.append(
                _build_import_row(
                    row_no=row_no,
                    sheet_name=sheet_name,
                    work_date=work_date,
                    afm=afm,
                    eponymo=str(emp.get("eponymo") or sched_ep or "").strip() or None,
                    onoma=str(emp.get("onoma") or sched_on or "").strip() or None,
                    import_action="absent",
                    intervals=[],
                    schedule_type="ΑΝ",
                    current_schedule=current_schedule,
                    row_errors=[],
                )
            )

    work_dates_unique = sorted(set(work_dates), key=lambda d: datetime.strptime(d, "%d/%m/%Y"))
    actual_label, actual_from, actual_to = _week_label_from_dates(work_dates_unique)
    if actual_label:
        instructions_label = str(meta.get("instructions_week_label") or meta.get("week_label") or "").strip()
        meta["week_label"] = actual_label
        meta["week_from"] = actual_from
        meta["week_to"] = actual_to
        if instructions_label and instructions_label != actual_label:
            errors.append(
                "Η εβδομάδα στο φύλλο «Οδηγίες» "
                f"({instructions_label}) διαφέρει από τις ημερομηνίες των φύλλων "
                f"({actual_label}) — χρησιμοποιούνται οι πραγματικές ημερομηνίες."
            )

    summary = summarize_import_rows(parsed_rows)
    return {
        "meta": meta,
        "work_dates": work_dates_unique,
        "rows": parsed_rows,
        "errors": errors,
        "summary": summary,
    }


def parse_weekly_schedule_workbook(
    file_bytes: bytes,
    *,
    employer_afm: str,
    branch_aa: str,
) -> dict[str, Any]:
    wb = load_workbook(filename=__import__("io").BytesIO(file_bytes), data_only=True)
    if INSTRUCTIONS_SHEET not in wb.sheetnames:
        raise ValueError("Το αρχείο δεν περιέχει φύλλο «Οδηγίες»")

    meta = _read_instructions(wb[INSTRUCTIONS_SHEET])
    employees = {
        norm_afm(emp.get("afm") or ""): emp
        for emp in list_employees_for_employer(employer_afm, branch_aa)
    }
    known_afms = set(employees.keys())

    if WEEK_SHEET in wb.sheetnames:
        return _parse_single_week_sheet(
            wb,
            meta=meta,
            employer_afm=employer_afm,
            branch_aa=branch_aa,
            employees=employees,
            known_afms=known_afms,
        )

    return _parse_legacy_multi_sheet(
        wb,
        meta=meta,
        employer_afm=employer_afm,
        branch_aa=branch_aa,
        employees=employees,
        known_afms=known_afms,
    )


def summarize_import_rows(rows: list[dict[str, Any]]) -> dict[str, int]:
    counts = {
        "total": 0,
        "skip": 0,
        "same": 0,
        "new": 0,
        "update": 0,
        "error": 0,
        "apply": 0,
        "absent": 0,
    }
    for row in rows:
        counts["total"] += 1
        kind = str(row.get("change_kind") or "").strip()
        if kind in counts:
            counts[kind] += 1
        if str(row.get("import_action") or "") == "absent":
            counts["absent"] += 1
        if kind in ("new", "update") and not row.get("validation_errors"):
            counts["apply"] += 1
    return counts


def format_snapshot_label(snapshot: list[dict[str, Any]]) -> str:
    if not snapshot:
        return "—"
    first = snapshot[0]
    st = str(first.get("schedule_type") or first.get("shift_type") or "").strip().upper()
    if st in ("ΑΝ", "AN") or re.search(r"ρεπο|ανάπαυση", st, re.I):
        return "ΡΕΠΟ"
    parts: list[str] = []
    for item in snapshot:
        hf = _hm_short(str(item.get("hour_from") or "")) or "—"
        ht = _hm_short(str(item.get("hour_to") or "")) or "—"
        parts.append(f"{hf}–{ht}")
    return " / ".join(parts)


def json_dumps(data: Any) -> str:
    return json.dumps(data, ensure_ascii=False)
