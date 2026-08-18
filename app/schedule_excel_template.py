"""Δημιουργία Excel template εβδομαδιαίου ωραρίου (κενές ώρες για συμπλήρωση)."""

from __future__ import annotations

import re
from datetime import date, timedelta
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from app.repo_entities import list_employees_for_employer
from app.repo_schedule import list_schedule_for_range
from app.repo_store import get_store_config
from app.schedule_excel_layout import (
    BASE_COL_COUNT,
    BASE_HEADERS,
    DAY_COUNT,
    DAY_FIELD_COUNT,
    DAY_FIELD_HEADERS,
    HOURS_COL,
    INSTRUCTIONS_SHEET,
    ROWS_PER_EMPLOYEE,
    SINGLE_SHEET_DATA_START_ROW,
    SINGLE_SHEET_HEADER_ROW_DAY,
    SINGLE_SHEET_HEADER_ROW_FIELDS,
    WEEK_SHEET,
    employee_block_start_row,
    single_sheet_day_col,
    single_sheet_last_col,
    weekly_hours_excel_formula,
)
from app.work_card_payload import norm_afm

DAYS = [
    ("Δευτέρα", 0),
    ("Τρίτη", 1),
    ("Τετάρτη", 2),
    ("Πέμπτη", 3),
    ("Παρασκευή", 4),
    ("Σάββατο", 5),
    ("Κυριακή", 6),
]


def resolve_week_monday(which: str, *, today: date | None = None) -> date:
    ref = today or date.today()
    monday = ref - timedelta(days=ref.weekday())
    key = str(which or "").strip().lower()
    if key == "next":
        return monday + timedelta(days=7)
    if key == "current":
        return monday
    raise ValueError("Μη έγκυρη εβδομάδα — επιτρέπονται current ή next")


def _safe_filename_part(value: str, *, fallback: str = "store") -> str:
    text = re.sub(r"[^\w\-]+", "_", str(value or "").strip(), flags=re.UNICODE)
    text = re.sub(r"_+", "_", text).strip("_")
    return (text[:48] or fallback)


def build_weekly_schedule_template_bytes(
    *,
    store_id: int,
    week_monday: date,
) -> tuple[bytes, str, dict[str, str]]:
    store = get_store_config(int(store_id))
    if not store:
        raise ValueError(f"Δεν βρέθηκε κατάστημα id={store_id}")

    employees = list_employees_for_employer(store["employer_afm"], store["branch_aa"])
    sunday = week_monday + timedelta(days=6)
    last_col = single_sheet_last_col()
    last_letter = get_column_letter(last_col)

    thin = Side(style="thin", color="D9E2F2")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    day_fill = PatternFill("solid", fgColor="2E75B6")
    warn_fill = PatternFill("solid", fgColor="FFF2CC")
    split_fill = PatternFill("solid", fgColor="F8FAFC")
    header_font = Font(color="FFFFFF", bold=True)
    day_font = Font(color="FFFFFF", bold=True, size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    wb = Workbook()
    info = wb.active
    info.title = INSTRUCTIONS_SHEET
    info["A1"] = f"Εβδομαδιαίο ωράριο - {store['name']}"
    info["A1"].font = Font(bold=True, size=14)
    info["A2"] = f"Εβδομάδα {week_monday.strftime('%d/%m/%Y')} - {sunday.strftime('%d/%m/%Y')}"
    info["A2"].font = Font(bold=True, size=11)
    instructions = [
        "",
        "Οδηγίες:",
        "1. Όλες οι ημέρες της εβδομάδας είναι στο φύλλο «Εβδομάδα».",
        "2. Κάθε εργαζόμενος έχει 2 γραμμές (συνενωμένες ΑΦΜ/Επώνυμο/Όνομα/Ώρες).",
        "   Η στήλη Ώρες υπολογίζει αυτόματα το σύνολο δηλωμένων ωρών της εβδομάδας.",
        "3. Ανά ημέρα: Ενέργεια, Από, Έως.",
        "4. Πάνω γραμμή = 1ο διάστημα (Από/Έως). Κάτω γραμμή = 2ο διάστημα (σπαστό).",
        "5. Οι ώρες/ΡΕΠΟ ξεκινούν κενά — συμπληρώστε μόνο ό,τι θέλετε να αλλάξει.",
        "6. Στήλη Ενέργεια: μόνο ΡΕΠΟ ή κενό.",
        "   - ΡΕΠΟ = ρεπό εκείνη την ημέρα (οι ώρες αγνοούνται).",
        "   - Κενό + ώρες συμπληρωμένες = αλλαγή / επιβεβαίωση ωραρίου.",
        "   - Κενό + κενές ώρες = καμία αλλαγή για την ημέρα.",
        "7. Οι ώρες: γράψε 4 ψηφία χωρίς άνω κάτω τελεία (π.χ. 0900, 1340).",
        "   Εμφανίζονται αυτόματα ως 09:00, 13:40.",
    ]
    r = 3
    for text in instructions:
        info[f"A{r}"] = text
        r += 1
    info["A17"] = "Store ID"
    info["B17"] = int(store_id)
    info["A18"] = "Employer AFM"
    info["B18"] = store["employer_afm"]
    info["A19"] = "Branch AA"
    info["B19"] = store["branch_aa"]
    info.column_dimensions["A"].width = 44
    info.column_dimensions["B"].width = 24

    ws = wb.create_sheet(WEEK_SHEET)
    ws["A1"] = (
        f"Εβδομαδιαίο ωράριο — {week_monday.strftime('%d/%m/%Y')} έως "
        f"{sunday.strftime('%d/%m/%Y')}"
    )
    ws["A1"].font = Font(bold=True, size=12)
    ws.merge_cells(f"A1:{last_letter}1")
    ws["A2"] = (
        "Κενό template για συμπλήρωση  |  "
        "Στήλη Ώρες = αυτόματο άθροισμα ωραρίου  |  "
        "Ενέργεια: ΡΕΠΟ ή κενό  |  Ώρες ημέρας: 4 ψηφία (0900→09:00)  |  "
        "Σπαστό: κάτω γραμμή = 2ο διάστημα (Από/Έως)"
    )
    ws["A2"].font = Font(italic=True)
    ws["A2"].fill = warn_fill
    ws.merge_cells(f"A2:{last_letter}2")

    for day_idx, (day_name, offset) in enumerate(DAYS):
        d = week_monday + timedelta(days=offset)
        start_col = single_sheet_day_col(day_idx, 0)
        end_col = single_sheet_day_col(day_idx, DAY_FIELD_COUNT - 1)
        cell = ws.cell(
            row=SINGLE_SHEET_HEADER_ROW_DAY,
            column=start_col,
            value=f"{day_name}\n{d.strftime('%d/%m/%Y')}",
        )
        cell.fill = day_fill
        cell.font = day_font
        cell.alignment = center
        cell.border = border
        ws.merge_cells(
            f"{get_column_letter(start_col)}{SINGLE_SHEET_HEADER_ROW_DAY}:"
            f"{get_column_letter(end_col)}{SINGLE_SHEET_HEADER_ROW_DAY}"
        )
    ws.row_dimensions[SINGLE_SHEET_HEADER_ROW_DAY].height = 30

    for c_idx, title in enumerate(BASE_HEADERS, start=1):
        cell = ws.cell(row=SINGLE_SHEET_HEADER_ROW_FIELDS, column=c_idx, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    for day_idx in range(DAY_COUNT):
        for f_idx, title in enumerate(DAY_FIELD_HEADERS):
            col = single_sheet_day_col(day_idx, f_idx)
            cell = ws.cell(row=SINGLE_SHEET_HEADER_ROW_FIELDS, column=col, value=title)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border

    work_dates = [
        (week_monday + timedelta(days=i)).strftime("%d/%m/%Y") for i in range(DAY_COUNT)
    ]
    _ = list_schedule_for_range(
        store["employer_afm"], store["branch_aa"], work_dates
    )

    hours_fill = PatternFill("solid", fgColor="D6EAF8")
    hours_font = Font(bold=True)
    start_row = SINGLE_SHEET_DATA_START_ROW
    for emp_idx, emp in enumerate(employees):
        r1 = employee_block_start_row(emp_idx)
        r2 = r1 + 1
        afm = norm_afm(str(emp.get("afm") or ""))
        ws.cell(row=r1, column=1, value=afm or str(emp.get("afm") or ""))
        ws.cell(row=r1, column=2, value=str(emp.get("eponymo") or ""))
        ws.cell(row=r1, column=3, value=str(emp.get("onoma") or ""))
        ws.cell(row=r1, column=HOURS_COL, value=weekly_hours_excel_formula(emp_idx))
        for c in range(1, BASE_COL_COUNT + 1):
            ws.merge_cells(
                start_row=r1, start_column=c, end_row=r2, end_column=c
            )
            for rr in (r1, r2):
                cell = ws.cell(row=rr, column=c)
                cell.border = border
                cell.alignment = center
                if c == HOURS_COL:
                    cell.fill = hours_fill
                    cell.font = hours_font
                    cell.number_format = "[h]:mm"

        for day_idx in range(DAY_COUNT):
            energia_col = single_sheet_day_col(day_idx, 0)
            ws.merge_cells(
                start_row=r1,
                start_column=energia_col,
                end_row=r2,
                end_column=energia_col,
            )
            for rr in (r1, r2):
                for f_idx in range(DAY_FIELD_COUNT):
                    col = single_sheet_day_col(day_idx, f_idx)
                    cell = ws.cell(row=rr, column=col)
                    cell.border = border
                    cell.alignment = center
                    if f_idx in (1, 2):
                        cell.number_format = r"00\:00"
                    if rr == r2 and f_idx in (1, 2):
                        cell.fill = split_fill

    max_row = (
        employee_block_start_row(len(employees) - 1) + ROWS_PER_EMPLOYEE - 1
        if employees
        else start_row
    )

    for day_idx in range(DAY_COUNT):
        energia_col = get_column_letter(single_sheet_day_col(day_idx, 0))
        dv_action = DataValidation(
            type="list", formula1='"ΡΕΠΟ"', allow_blank=True, showDropDown=False
        )
        dv_action.errorTitle = "Μη έγκυρη τιμή"
        dv_action.error = "Επιτρέπεται μόνο ΡΕΠΟ ή κενό."
        ws.add_data_validation(dv_action)
        if employees:
            for emp_idx in range(len(employees)):
                r1 = employee_block_start_row(emp_idx)
                dv_action.add(f"{energia_col}{r1}")

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 10
    for day_idx in range(DAY_COUNT):
        for f_idx, width in enumerate((11, 9, 9)):
            col = get_column_letter(single_sheet_day_col(day_idx, f_idx))
            ws.column_dimensions[col].width = width

    ws.freeze_panes = f"A{start_row}"
    if employees:
        ws.auto_filter.ref = f"A{SINGLE_SHEET_HEADER_ROW_FIELDS}:{last_letter}{max_row}"

    bio = BytesIO()
    wb.save(bio)
    store_part = _safe_filename_part(str(store.get("name") or f"store_{store_id}"))
    filename = (
        f"weekly_schedule_{store_part}_"
        f"{week_monday.strftime('%Y%m%d')}_{sunday.strftime('%Y%m%d')}.xlsx"
    )
    meta = {
        "week_from": week_monday.strftime("%d/%m/%Y"),
        "week_to": sunday.strftime("%d/%m/%Y"),
        "employee_count": str(len(employees)),
        "filled_day_slots": "0",
    }
    return bio.getvalue(), filename, meta
