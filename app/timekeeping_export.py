"""Excel export for calculated retrospective timekeeping."""

from __future__ import annotations

from datetime import datetime
from io import BytesIO
import re
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


_NAVY = "17324D"
_BLUE = "2F75B5"
_LIGHT = "EAF2F8"
_BORDER = Side(style="thin", color="D9E2E9")
_BREAKDOWN_KEYS = ("day", "night", "sunday_holiday", "night_sunday_holiday")
_BREAKDOWN_LABELS = ("Ημέρας", "Νύχτας", "Κυρ/Αργίας", "Νύχτας/Κυρ-Αργίας")
_HOURS_RE = re.compile(r"\d{1,2}:\d{2}\s*[–-]\s*\d{1,2}:\d{2}")


def _duration(minutes: Any) -> float:
    return max(0, int(minutes or 0)) / 60


def _breakdown_values(item: dict[str, Any], field: str) -> list[float]:
    values = item.get(field) or {}
    return [_duration(values.get(key)) for key in _BREAKDOWN_KEYS]


def _family_headers(prefix: str) -> list[str]:
    return [f"{prefix} – {label} (ώρες)" for label in _BREAKDOWN_LABELS]


def _hours_only(value: Any) -> str:
    text = str(value or "").strip()
    return text if _HOURS_RE.search(text) else ""


def _style_sheet(ws, *, title: str, meta: str, headers: list[str], widths: list[int]) -> int:
    ws.sheet_view.showGridLines = False
    ws.append([title])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws["A1"].font = Font(name="Aptos Display", size=18, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=_NAVY)
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 30
    ws.append([meta])
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws["A2"].font = Font(name="Aptos", size=10, color="526777")
    ws.append(headers)
    header_row = 3
    for cell in ws[header_row]:
        cell.font = Font(name="Aptos", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=_BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=_BORDER)
    ws.row_dimensions[header_row].height = 34
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.freeze_panes = "A4"
    return header_row


def _finish_table(ws, header_row: int, duration_from: int, duration_to: int) -> None:
    if ws.max_row > header_row:
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ws.max_column)}{ws.max_row}"
    for row in range(header_row + 1, ws.max_row + 1):
        if row % 2 == 0:
            for cell in ws[row]:
                cell.fill = PatternFill("solid", fgColor=_LIGHT)
        for col in range(duration_from, duration_to + 1):
            ws.cell(row, col).number_format = "0.##"
            ws.cell(row, col).alignment = Alignment(horizontal="right")
        for cell in ws[row]:
            cell.border = Border(bottom=_BORDER)
            cell.font = Font(name="Aptos", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=cell.column in (1, ws.max_column))


def build_timekeeping_export_xlsx(
    *, report: dict[str, Any], meta_line: str,
    title: str = "Ωρομέτρηση εβδομάδας",
    daily_title: str = "Αναλυτική ωρομέτρηση ανά ημέρα",
) -> bytes:
    wb = Workbook()
    families = (
        ("Υπερεργασία 20%", "overwork_breakdown"),
        ("Πρόσθετη μερικής 12%", "partial_additional_12_breakdown"),
        ("Υπερωρία 40%", "overtime_40_breakdown"),
        ("Υπερωρία 60%", "overtime_60_breakdown"),
        ("Κατ’ εξαίρεση", "overtime_120_breakdown"),
        ("6η ημέρα 30%", "sixth_day_breakdown"),
        ("6η ημέρα άνω των 48 ωρών", "sixth_day_above_48_breakdown"),
        ("Κατ’ εξαίρεση 6η ημέρα άνω των 48 ωρών", "exception_sixth_day_above_48_breakdown"),
    )

    summary = wb.active
    summary.title = "Σύνοψη"
    summary_headers = ["Εργαζόμενος", "ΑΦΜ", "Αναγνωρισμένη βάση (ώρες)"] + _family_headers("Βάση")
    for label, _ in families:
        summary_headers += _family_headers(label)
    summary_headers += ["Ετήσιες νόμιμες υπερωρίες μετά την περίοδο"]
    header_row = _style_sheet(
        summary, title=title, meta=meta_line, headers=summary_headers,
        widths=[28, 14, 20] + [18] * (4 + 4 * len(families)) + [25],
    )
    for item in report.get("employees") or []:
        values = [
            f"{item.get('eponymo') or ''} {item.get('onoma') or ''}".strip(),
            str(item.get("employee_afm") or ""),
            _duration(item.get("recognized_work_minutes")),
            _duration(item.get("day")), _duration(item.get("night")),
            _duration(item.get("sunday_holiday")), _duration(item.get("night_sunday_holiday")),
        ]
        for _, field in families:
            values += _breakdown_values(item, field)
        values.append(_duration(item.get("annual_legal_overtime_minutes_after_period")))
        summary.append(values)
    _finish_table(summary, header_row, 3, len(summary_headers))

    daily = wb.create_sheet("Ανά ημέρα")
    daily_headers = [
        "Ημερομηνία", "Εργαζόμενος", "ΑΦΜ", "Κατάσταση", "Πηγή βάσης",
        "Αναγνωρισμένο ωράριο", "Διάλειμμα", "Καθαρή βάση (ώρες)",
    ] + _family_headers("Βάση")
    for label, _ in families:
        daily_headers += _family_headers(label)
    daily_headers += ["Παρατηρήσεις"]
    daily_header = _style_sheet(
        daily, title=daily_title, meta=meta_line, headers=daily_headers,
        widths=[14, 28, 14, 12, 22, 24, 18, 16] + [18] * (4 + 4 * len(families)) + [46],
    )
    for item in report.get("days") or []:
        premiums = item.get("premium_minutes") or {}
        values = [
            item.get("work_date") or "",
            f"{item.get('eponymo') or ''} {item.get('onoma') or ''}".strip(),
            str(item.get("employee_afm") or ""), item.get("status") or "",
            item.get("basis_source") or "", item.get("basis_label") or "",
            item.get("break_interval") or "", _duration(item.get("recognized_work_minutes")),
            *_breakdown_values({"base": premiums}, "base"),
        ]
        for _, field in families:
            values += _breakdown_values(item, field)
        values.append(" · ".join(str(value) for value in item.get("warnings") or []))
        daily.append(values)
    _finish_table(daily, daily_header, 8, len(daily_headers) - 1)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_timekeeping_detailed_export_xlsx(
    *, report: dict[str, Any], store: dict[str, Any], meta_line: str,
    title: str = "Πλήρης ανάλυση ωρομέτρησης ανά εργαζόμενο",
) -> bytes:
    """Render the common timekeeping report as one auditable employee/day table.

    This function deliberately performs no payroll classification.  Every
    duration comes from ``build_timekeeping_report`` so summary and detailed
    exports cannot drift into separate rule engines.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Πλήρης ανάλυση"
    ws.sheet_view.showGridLines = False

    headers = [
        "Εργοδότης ΑΦΜ", "Κωδικός υποκ/τος", "Υποκατάστημα",
        "Επώνυμο", "Όνομα", "ΑΦΜ εργαζομένου", "Μερική απασχόληση",
        "Ημέρα", "Ημερομηνία", "Δηλωμένο / προτεινόμενο ωράριο",
        "Αναγνωρισμένο ωράριο",
        "Κινήσεις κάρτας", "Μικτή διάρκεια (ώρες)", "Διάλειμμα", "Καθαρή βάση (ώρες)",
        "Ώρες ημέρας", "Ώρες νύχτας 25%", "Ώρες Κυρ/Αργίας 75%",
        "Ώρες νύχτας + Κυρ/Αργίας",
    ]
    headers += _family_headers("Υπερεργασία 20%")
    headers += _family_headers("Πρόσθετη μερικής 12%") + ["Διάστημα πρόσθετης μερικής"]
    headers += _family_headers("Υπερωρία 40%")
    headers += _family_headers("Υπερωρία 60%")
    headers += _family_headers("Κατ’ εξαίρεση")
    headers += _family_headers("6η ημέρα 30%") + ["Σύνολο 6ης ημέρας"]
    headers += _family_headers("6η ημέρα άνω των 48 ωρών") + ["Σύνολο 6ης άνω των 48"]
    headers += _family_headers("Κατ’ εξαίρεση 6η ημέρα άνω των 48 ωρών") + ["Σύνολο κατ’ εξαίρεση 6ης άνω των 48"]
    headers += ["Κατάσταση ημέρας", "Πηγή βάσης", "Παρατηρήσεις"]
    groups = [
        (1, 3, "Επιχείρηση"), (4, 7, "Εργαζόμενος"), (8, 15, "Ημερήσια στοιχεία"),
        (16, 19, "Προσαυξήσεις βάσης"), (20, 23, "Υπερεργασία 20%"),
        (24, 28, "Πρόσθετη μερικής 12%"), (29, 32, "Υπερωρία 40%"),
        (33, 36, "Υπερωρία 60%"), (37, 40, "Κατ’ εξαίρεση"),
        (41, 45, "6η ημέρα 30%"), (46, 50, "6η ημέρα άνω των 48 ωρών"),
        (51, 55, "Κατ’ εξαίρεση 6η ημέρα άνω των 48 ωρών"),
        (56, 58, "Έλεγχος"),
    ]
    for start, end, label in groups:
        ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
        cell = ws.cell(1, start, label)
        cell.fill = PatternFill("solid", fgColor=_NAVY)
        cell.font = Font(name="Aptos", size=10, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        for col in range(start, end + 1):
            ws.cell(1, col).fill = PatternFill("solid", fgColor=_NAVY)
    ws.row_dimensions[1].height = 24
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws.cell(2, 1, title)
    ws.cell(2, 1).font = Font(name="Aptos Display", size=16, bold=True, color=_NAVY)
    ws.cell(2, 1).alignment = Alignment(vertical="center")
    ws.row_dimensions[2].height = 28
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(headers))
    ws.cell(3, 1, meta_line)
    ws.cell(3, 1).font = Font(name="Aptos", size=10, color="526777")
    ws.append(headers)
    for cell in ws[4]:
        cell.fill = PatternFill("solid", fgColor=_BLUE)
        cell.font = Font(name="Aptos", size=9, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=_BORDER)
    ws.row_dimensions[4].height = 44

    weekday_names = ("Δευ", "Τρι", "Τετ", "Πεμ", "Παρ", "Σαβ", "Κυρ")
    for item in report.get("days") or []:
        premiums = item.get("premium_minutes") or {}
        work_date = datetime.strptime(str(item.get("work_date") or ""), "%d/%m/%Y").date()
        ws.append([
            str(store.get("employer_afm") or ""), str(store.get("branch_aa") or ""),
            str(store.get("name") or ""), item.get("eponymo") or "", item.get("onoma") or "",
            str(item.get("employee_afm") or ""), "ΝΑΙ" if str(item.get("contract_kind") or "") == "Μερική" else "ΟΧΙ",
            weekday_names[work_date.weekday()], work_date,
            _hours_only(item.get("effective_declared") or item.get("declared")),
            item.get("basis_label") or "", item.get("punch_recorded") or item.get("actual") or "",
            _duration(item.get("recognized_span_minutes")), item.get("break_interval") or "",
            _duration(item.get("recognized_work_minutes")), _duration(premiums.get("day")),
            _duration(premiums.get("night")), _duration(premiums.get("sunday_holiday")),
            _duration(premiums.get("night_sunday_holiday")),
            *_breakdown_values(item, "overwork_breakdown"),
            *_breakdown_values(item, "partial_additional_12_breakdown"),
            " · ".join(str(value) for value in item.get("partial_additional_12_intervals") or []),
            *_breakdown_values(item, "overtime_40_breakdown"),
            *_breakdown_values(item, "overtime_60_breakdown"),
            *_breakdown_values(item, "overtime_120_breakdown"),
            *_breakdown_values(item, "sixth_day_breakdown"),
            _duration(item.get("sixth_day_minutes")),
            *_breakdown_values(item, "sixth_day_above_48_breakdown"),
            _duration(item.get("sixth_day_above_48_minutes")),
            *_breakdown_values(item, "exception_sixth_day_above_48_breakdown"),
            _duration(item.get("exception_sixth_day_above_48_minutes")),
            item.get("day_state") or "",
            item.get("basis_source") or "", " · ".join(str(value) for value in item.get("warnings") or []),
        ])

    widths = [15, 15, 24, 22, 18, 16, 16, 10, 13, 30, 24, 25, 15, 18, 15,
              15, 17, 19, 23] + [18] * 8 + [24] + [18] * 24 + [18, 18, 18, 20, 42]
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.freeze_panes = "J5"
    if ws.max_row > 4:
        ws.auto_filter.ref = f"A4:{get_column_letter(len(headers))}{ws.max_row}"
    duration_columns = {13} | set(range(15, 28)) | set(range(29, 56))
    for row in range(5, ws.max_row + 1):
        if row % 2:
            for cell in ws[row]:
                cell.fill = PatternFill("solid", fgColor=_LIGHT)
        ws.cell(row, 9).number_format = "dd/mm/yyyy"
        for col in duration_columns:
            ws.cell(row, col).number_format = "0.##"
            ws.cell(row, col).alignment = Alignment(horizontal="right", vertical="top")
        for cell in ws[row]:
            cell.border = Border(bottom=_BORDER)
            cell.font = Font(name="Aptos", size=9)
            if cell.column not in duration_columns:
                cell.alignment = Alignment(vertical="top", wrap_text=cell.column in (10, 11, 12, 28, 48))
    ws.auto_filter.ref = f"A4:{get_column_letter(len(headers))}{ws.max_row}"
    ws.print_title_rows = "1:4"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
