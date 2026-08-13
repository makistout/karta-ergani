"""Εξαγωγή ορατών αποτελεσμάτων απολογιστικού σε Excel."""

from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def _safe_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\r\n", "\n").replace("\r", "\n").strip()


def build_apologistic_export_xlsx(
    *,
    meta_line: str,
    headers: list[str],
    rows: list[list[Any]],
) -> bytes:
    if not headers:
        raise ValueError("Λείπουν κεφαλίδες εξαγωγής")
    if not rows:
        raise ValueError("Δεν υπάρχουν γραμμές για εξαγωγή")

    wb = Workbook()
    ws = wb.active
    ws.title = "Απολογιστικό"

    ws.append([_safe_cell(meta_line)])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws["A1"].font = Font(bold=True)

    ws.append([_safe_cell(item) for item in headers])
    header_row = ws.max_row
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.font = Font(bold=True)

    for row in rows:
        ws.append([_safe_cell(item) for item in row])

    for col_idx, header in enumerate(headers, start=1):
        letter = get_column_letter(col_idx)
        max_len = len(_safe_cell(header))
        for row_idx in range(header_row + 1, ws.max_row + 1):
            max_len = max(max_len, len(_safe_cell(ws.cell(row=row_idx, column=col_idx).value)))
        ws.column_dimensions[letter].width = min(max(10, max_len + 2), 48)

    ws.freeze_panes = f"A{header_row + 1}"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
