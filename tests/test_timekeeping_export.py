from io import BytesIO
from openpyxl import load_workbook

from app.timekeeping_export import (
    build_timekeeping_detailed_export_xlsx,
    build_timekeeping_export_xlsx,
)


def test_timekeeping_export_has_summary_and_daily_sheets_with_typed_durations():
    report = {
        "employees": [{
            "employee_afm": "012345678", "eponymo": "ΔΟΚΙΜΗ", "onoma": "ΕΝΑ",
            "recognized_work_minutes": 480, "day": 420, "night": 60,
            "sunday_holiday": 0, "night_sunday_holiday": 0,
            "partial_additional_12": 0, "sixth_day_minutes": 0,
            "overtime_40": 30, "overtime_60": 0, "overtime_120": 0,
            "annual_legal_overtime_minutes_after_period": 300,
            "overtime_40_breakdown": {"day": 30, "night": 0, "sunday_holiday": 0, "night_sunday_holiday": 0},
        }],
        "days": [{
            "work_date": "03/08/2026", "employee_afm": "012345678",
            "eponymo": "ΔΟΚΙΜΗ", "onoma": "ΕΝΑ", "status": "ok",
            "basis_source": "declared_compliant", "basis_label": "14:00–22:00",
            "break_interval": "", "recognized_work_minutes": 480,
            "premium_minutes": {"day": 420, "night": 60, "sunday_holiday": 0, "night_sunday_holiday": 0},
            "partial_additional_12": 0, "sixth_day_minutes": 0,
            "overtime_40": 30, "overtime_60": 0, "overtime_120": 0,
            "overtime_40_breakdown": {"day": 30, "night": 0, "sunday_holiday": 0, "night_sunday_holiday": 0},
            "warnings": [],
        }],
    }
    content = build_timekeeping_export_xlsx(report=report, meta_line="Store · 03/08/2026–09/08/2026")
    workbook = load_workbook(BytesIO(content), data_only=False)
    assert workbook.sheetnames == ["Σύνοψη", "Ανά ημέρα"]
    assert workbook["Σύνοψη"]["B4"].value == "012345678"
    assert workbook["Σύνοψη"]["C4"].value == 8
    assert workbook["Σύνοψη"]["C4"].number_format == "0.##"
    assert workbook["Σύνοψη"]["P3"].value == "Υπερωρία 40% – Ημέρας (ώρες)"
    assert workbook["Σύνοψη"]["P4"].value == 0.5
    assert workbook["Ανά ημέρα"]["F4"].value == "14:00–22:00"


def test_detailed_export_projects_common_daily_report_without_recalculation():
    report = {"days": [{
        "employee_afm": "012345678", "eponymo": "ΔΟΚΙΜΗ", "onoma": "ΕΝΑ",
        "work_date": "17/08/2026", "contract_kind": "Μερική", "status": "change",
        "declared": "09:00–13:00", "effective_declared": "09:00–15:00",
        "basis_label": "09:00–15:00",
        "punch_recorded": "09:01–15:02", "recognized_span_minutes": 360,
        "recognized_work_minutes": 360, "break_interval": "",
        "premium_minutes": {"day": 360, "night": 0, "sunday_holiday": 0, "night_sunday_holiday": 0},
        "overwork_minutes": 0, "partial_additional_12": 120,
        "partial_additional_12_intervals": ["13:00–15:00"],
        "overwork_breakdown": {"day": 0, "night": 0, "sunday_holiday": 0, "night_sunday_holiday": 0},
        "partial_additional_12_breakdown": {"day": 120, "night": 0, "sunday_holiday": 0, "night_sunday_holiday": 0},
        "overtime_40": 0, "overtime_60": 0, "overtime_120": 0,
        "overtime_40_breakdown": {"day": 0, "night": 0, "sunday_holiday": 0, "night_sunday_holiday": 0},
        "overtime_60_breakdown": {"day": 0, "night": 0, "sunday_holiday": 0, "night_sunday_holiday": 0},
        "overtime_120_breakdown": {"day": 0, "night": 0, "sunday_holiday": 0, "night_sunday_holiday": 0},
        "sixth_day_minutes": 0, "day_state": "Εργασία", "basis_source": "effective_proposed",
        "sixth_day_breakdown": {"day": 0, "night": 0, "sunday_holiday": 0, "night_sunday_holiday": 0},
        "warnings": [],
    }]}
    content = build_timekeeping_detailed_export_xlsx(
        report=report,
        store={"name": "Store", "employer_afm": "111111111", "branch_aa": "0"},
        meta_line="Store · 17/08/2026–23/08/2026",
    )
    workbook = load_workbook(BytesIO(content), data_only=False)
    sheet = workbook["Πλήρης ανάλυση"]
    assert sheet["F5"].value == "012345678"
    assert sheet["I5"].value.strftime("%d/%m/%Y") == "17/08/2026"
    assert sheet["J5"].value == "09:00–15:00"
    assert sheet["K5"].value == "09:00–15:00"
    assert sheet["O5"].value == 6
    assert sheet["X5"].value == 2
    assert sheet["X5"].number_format == "0.##"
    assert sheet["AB5"].value == "13:00–15:00"
