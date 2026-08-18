import unittest
from datetime import date
from io import BytesIO
from unittest.mock import patch

from openpyxl import load_workbook

from app.schedule_excel_import import (
    _afms_missing_from_sheet,
    _build_import_row,
    _format_time_cell,
    parse_weekly_schedule_workbook,
    summarize_import_rows,
)
from app.schedule_excel_layout import WEEK_SHEET
from app.schedule_excel_template import build_weekly_schedule_template_bytes, resolve_week_monday


class ScheduleExcelTemplateTests(unittest.TestCase):
    def test_resolve_week_monday_current_and_next(self):
        ref = date(2026, 7, 7)  # Tuesday
        self.assertEqual(resolve_week_monday("current", today=ref), date(2026, 7, 6))
        self.assertEqual(resolve_week_monday("next", today=ref), date(2026, 7, 13))


class ScheduleExcelImportAbsentTests(unittest.TestCase):
    def test_afms_missing_from_sheet_includes_known_and_scheduled(self):
        missing = _afms_missing_from_sheet(
            sheet_afms={"111111111"},
            known_afms={"111111111", "222222222"},
            current_schedule=[
                {"employee_afm": "333333333", "hour_from": "09:00", "hour_to": "17:00"},
            ],
        )
        self.assertEqual(missing, {"222222222", "333333333"})

    def test_absent_row_marks_working_employee_for_rest_update(self):
        current = [
            {
                "employee_afm": "162094518",
                "eponymo": "ΜΠΟΓΡΗΣ",
                "onoma": "ΓΕΩΡΓΙΟΣ",
                "hour_from": "11:00",
                "hour_to": "19:00",
                "shift_type": "ΕΡΓΑΣΙΑ",
            }
        ]
        row = _build_import_row(
            row_no=1,
            sheet_name="Τετάρτη 08-07",
            work_date="08/07/2026",
            afm="162094518",
            eponymo="ΜΠΟΓΡΗΣ",
            onoma="ΓΕΩΡΓΙΟΣ",
            import_action="absent",
            intervals=[],
            schedule_type="ΑΝ",
            current_schedule=current,
        )
        self.assertEqual(row["import_action"], "absent")
        self.assertEqual(row["change_kind"], "update")
        self.assertEqual(row["proposed_label"] if "proposed_label" in row else None, None)

    def test_absent_row_skips_when_already_rest(self):
        current = [
            {
                "employee_afm": "157216372",
                "eponymo": "ΠΡΩΤΟΥΛΗΣ",
                "shift_type": "ΑΝΑΠΑΥΣΗ/ΡΕΠΟ",
            }
        ]
        row = _build_import_row(
            row_no=1,
            sheet_name="Τετάρτη 08-07",
            work_date="08/07/2026",
            afm="157216372",
            eponymo="ΠΡΩΤΟΥΛΗΣ",
            onoma="ΒΑΣΙΛΕΙΟΣ",
            import_action="absent",
            intervals=[],
            schedule_type="ΑΝ",
            current_schedule=current,
        )
        self.assertEqual(row["change_kind"], "same")

    def test_summarize_counts_absent_rows(self):
        rows = [
            {"change_kind": "update", "import_action": "absent", "validation_errors": []},
            {"change_kind": "skip", "import_action": "skip", "validation_errors": []},
        ]
        summary = summarize_import_rows(rows)
        self.assertEqual(summary["absent"], 1)
        self.assertEqual(summary["apply"], 1)


class ScheduleExcelTimeCellTests(unittest.TestCase):
    def test_format_hhmm_digits(self):
        self.assertEqual(_format_time_cell(900), "09:00")
        self.assertEqual(_format_time_cell(1340), "13:40")
        self.assertEqual(_format_time_cell("0900"), "09:00")
        self.assertEqual(_format_time_cell("13:40"), "13:40")

    def test_format_excel_fraction_still_works(self):
        # 0.375 ημέρας = 09:00
        self.assertEqual(_format_time_cell(0.375), "09:00")

    def test_format_invalid_minutes(self):
        self.assertEqual(_format_time_cell(1399), "")
        self.assertEqual(_format_time_cell(2500), "")


class ScheduleExcelSingleSheetTests(unittest.TestCase):
    def test_single_sheet_template_and_parse_roundtrip(self):
        week_monday = date(2026, 7, 6)
        store = {
            "id": 1,
            "name": "Test Store",
            "employer_afm": "123456789",
            "branch_aa": "0",
        }
        employees = [
            {"afm": "111111111", "eponymo": "TEST", "onoma": "ONE"},
            {"afm": "222222222", "eponymo": "TEST", "onoma": "TWO"},
        ]

        with patch("app.schedule_excel_template.get_store_config", return_value=store), patch(
            "app.schedule_excel_template.list_employees_for_employer", return_value=employees
        ), patch("app.schedule_excel_template.list_schedule_for_range", return_value=[]):
            xlsx, _filename, _meta = build_weekly_schedule_template_bytes(
                store_id=1,
                week_monday=week_monday,
            )

        wb = load_workbook(filename=BytesIO(xlsx))
        self.assertIn(WEEK_SHEET, wb.sheetnames)
        self.assertEqual(len(wb.sheetnames), 2)

        from app.schedule_excel_layout import HOURS_COL, HOURS_HEADER, single_sheet_day_col

        ws = wb[WEEK_SHEET]
        self.assertEqual(ws.cell(row=4, column=HOURS_COL).value, HOURS_HEADER)
        hours_formula = str(ws.cell(row=5, column=HOURS_COL).value or "")
        self.assertTrue(hours_formula.startswith("="), hours_formula)
        self.assertIn("ΡΕΠΟ", hours_formula)
        # Πάνω γραμμή εργαζομένου 1 · Από Δευτέρας.
        self.assertEqual(
            ws.cell(row=5, column=single_sheet_day_col(0, 1)).number_format,
            r"00\:00",
        )
        self.assertEqual(ws.cell(row=5, column=single_sheet_day_col(0, 0)).value, None)
        # 2 γραμμές ανά εργαζόμενο: AFM μόνο στην πάνω (merged).
        self.assertEqual(str(ws.cell(row=5, column=1).value), "111111111")
        self.assertIsNone(ws.cell(row=6, column=1).value)

        # Δευτέρα ΡΕΠΟ · Τρίτη 09:00–17:00 · Τετάρτη σπαστό 10–14 + 17–21
        ws.cell(row=5, column=single_sheet_day_col(0, 0), value="ΡΕΠΟ")
        ws.cell(row=5, column=single_sheet_day_col(1, 1), value=900)
        ws.cell(row=5, column=single_sheet_day_col(1, 2), value=1700)
        ws.cell(row=5, column=single_sheet_day_col(2, 1), value=1000)
        ws.cell(row=5, column=single_sheet_day_col(2, 2), value=1400)
        ws.cell(row=6, column=single_sheet_day_col(2, 1), value=1700)
        ws.cell(row=6, column=single_sheet_day_col(2, 2), value=2100)
        bio = BytesIO()
        wb.save(bio)
        filled = bio.getvalue()

        with patch(
            "app.schedule_excel_import.list_employees_for_employer", return_value=employees
        ), patch("app.schedule_excel_import.list_schedule_for_store", return_value=[]):
            parsed = parse_weekly_schedule_workbook(
                filled,
                employer_afm="123456789",
                branch_aa="0",
            )

        self.assertEqual(parsed["work_dates"], [
            "06/07/2026",
            "07/07/2026",
            "08/07/2026",
            "09/07/2026",
            "10/07/2026",
            "11/07/2026",
            "12/07/2026",
        ])
        rows = parsed["rows"]
        self.assertEqual(len(rows), 14)
        mon_repo = [r for r in rows if r["employee_afm"] == "111111111" and r["work_date"] == "06/07/2026"][0]
        self.assertEqual(mon_repo["import_action"], "rest")
        tue_work = [r for r in rows if r["employee_afm"] == "111111111" and r["work_date"] == "07/07/2026"][0]
        self.assertEqual(tue_work["import_action"], "work")
        self.assertEqual(tue_work["hour_from_1"], "09:00")
        self.assertEqual(tue_work["hour_to_1"], "17:00")
        wed_split = [r for r in rows if r["employee_afm"] == "111111111" and r["work_date"] == "08/07/2026"][0]
        self.assertEqual(wed_split["hour_from_2"], "17:00")
        self.assertEqual(wed_split["hour_to_2"], "21:00")
        skips = [r for r in rows if r["import_action"] == "skip"]
        self.assertGreater(len(skips), 0)

    def test_template_is_blank_hours(self):
        from app.schedule_excel_layout import employee_block_start_row, single_sheet_day_col

        week_monday = date(2026, 7, 6)
        store = {
            "id": 1,
            "name": "Test Store",
            "employer_afm": "123456789",
            "branch_aa": "0",
        }
        employees = [
            {"afm": "111111111", "eponymo": "TEST", "onoma": "ONE"},
        ]

        with patch("app.schedule_excel_template.get_store_config", return_value=store), patch(
            "app.schedule_excel_template.list_employees_for_employer", return_value=employees
        ), patch("app.schedule_excel_template.list_schedule_for_range", return_value=[]):
            xlsx, _filename, meta = build_weekly_schedule_template_bytes(
                store_id=1,
                week_monday=week_monday,
            )

        self.assertEqual(meta.get("filled_day_slots"), "0")
        ws = load_workbook(filename=BytesIO(xlsx))[WEEK_SHEET]
        r1 = employee_block_start_row(0)
        r2 = r1 + 1
        self.assertEqual(ws.cell(row=r1, column=1).value, "111111111")
        for day_idx in range(7):
            self.assertIsNone(ws.cell(row=r1, column=single_sheet_day_col(day_idx, 0)).value)
            self.assertIsNone(ws.cell(row=r1, column=single_sheet_day_col(day_idx, 1)).value)
            self.assertIsNone(ws.cell(row=r1, column=single_sheet_day_col(day_idx, 2)).value)
            self.assertIsNone(ws.cell(row=r2, column=single_sheet_day_col(day_idx, 1)).value)
            self.assertIsNone(ws.cell(row=r2, column=single_sheet_day_col(day_idx, 2)).value)

    def test_hours_column_sums_declared_schedule(self):
        from app.schedule_excel_layout import HOURS_COL, format_hours_minutes, weekly_declared_minutes

        week_monday = date(2026, 7, 6)
        store = {
            "id": 1,
            "name": "Test Store",
            "employer_afm": "123456789",
            "branch_aa": "0",
        }
        employees = [
            {"afm": "111111111", "eponymo": "TEST", "onoma": "ONE"},
            {"afm": "222222222", "eponymo": "TEST", "onoma": "TWO"},
        ]
        schedule_rows = [
            {
                "employee_afm": "111111111",
                "work_date": "06/07/2026",
                "hour_from": "09:00",
                "hour_to": "17:00",
                "shift_type": "ΕΡΓΑΣΙΑ",
            },
            {
                "employee_afm": "111111111",
                "work_date": "07/07/2026",
                "hour_from": "10:00",
                "hour_to": "14:00",
                "shift_type": "ΕΡΓΑΣΙΑ",
            },
            {
                "employee_afm": "111111111",
                "work_date": "07/07/2026",
                "hour_from": "17:00",
                "hour_to": "21:00",
                "shift_type": "ΕΡΓΑΣΙΑ",
            },
            {
                "employee_afm": "111111111",
                "work_date": "08/07/2026",
                "hour_from": "22:00",
                "hour_to": "06:00",
                "shift_type": "ΕΡΓΑΣΙΑ",
            },
            {
                "employee_afm": "111111111",
                "work_date": "09/07/2026",
                "shift_type": "ΑΝΑΠΑΥΣΗ/ΡΕΠΟ",
            },
            {
                "employee_afm": "222222222",
                "work_date": "06/07/2026",
                "shift_type": "ADKAN",
            },
        ]

        with patch("app.schedule_excel_template.get_store_config", return_value=store), patch(
            "app.schedule_excel_template.list_employees_for_employer", return_value=employees
        ), patch(
            "app.schedule_excel_template.list_schedule_for_range", return_value=schedule_rows
        ):
            xlsx, _filename, _meta = build_weekly_schedule_template_bytes(
                store_id=1,
                week_monday=week_monday,
            )

        ws = load_workbook(filename=BytesIO(xlsx))[WEEK_SHEET]
        hours_formula = str(ws.cell(row=5, column=HOURS_COL).value or "")
        self.assertTrue(hours_formula.startswith("="))
        self.assertEqual(str(ws.cell(row=7, column=HOURS_COL).value or "").startswith("="), True)
        self.assertEqual(
            format_hours_minutes(weekly_declared_minutes(schedule_rows[:4])),
            "24:00",
        )

    def test_import_accepts_legacy_file_without_hours_column(self):
        from openpyxl import Workbook

        from app.schedule_excel_layout import (
            INSTRUCTIONS_SHEET,
            SINGLE_SHEET_DATA_START_ROW,
            SINGLE_SHEET_HEADER_ROW_FIELDS,
            WEEK_SHEET,
            single_sheet_day_col,
        )

        wb = Workbook()
        info = wb.active
        info.title = INSTRUCTIONS_SHEET
        info["A2"] = "06/07/2026 - 12/07/2026"
        ws = wb.create_sheet(WEEK_SHEET)
        ws.cell(row=SINGLE_SHEET_HEADER_ROW_FIELDS, column=1, value="ΑΦΜ")
        ws.cell(row=SINGLE_SHEET_HEADER_ROW_FIELDS, column=2, value="Επώνυμο")
        ws.cell(row=SINGLE_SHEET_HEADER_ROW_FIELDS, column=3, value="Όνομα")
        ws.cell(row=SINGLE_SHEET_HEADER_ROW_FIELDS, column=4, value="Ενέργεια")
        ws.cell(row=SINGLE_SHEET_HEADER_ROW_FIELDS, column=5, value="Από")
        ws.cell(row=SINGLE_SHEET_HEADER_ROW_FIELDS, column=6, value="Έως")
        r1 = SINGLE_SHEET_DATA_START_ROW
        ws.cell(row=r1, column=1, value="111111111")
        ws.cell(row=r1, column=2, value="TEST")
        ws.cell(row=r1, column=3, value="ONE")
        ws.cell(
            row=r1,
            column=single_sheet_day_col(0, 0, base_col_count=3),
            value="ΡΕΠΟ",
        )
        bio = BytesIO()
        wb.save(bio)

        employees = [{"afm": "111111111", "eponymo": "TEST", "onoma": "ONE"}]
        with patch(
            "app.schedule_excel_import.list_employees_for_employer", return_value=employees
        ), patch("app.schedule_excel_import.list_schedule_for_store", return_value=[]):
            parsed = parse_weekly_schedule_workbook(
                bio.getvalue(),
                employer_afm="123456789",
                branch_aa="0",
            )

        mon_repo = [
            r for r in parsed["rows"] if r["employee_afm"] == "111111111" and r["work_date"] == "06/07/2026"
        ][0]
        self.assertEqual(mon_repo["import_action"], "rest")
        header_errors = [e for e in parsed["errors"] if "headers" in e]
        self.assertEqual(header_errors, [])


class ScheduleExcelHoursCalcTests(unittest.TestCase):
    def test_weekly_hours_formula_sums_all_days(self):
        from app.schedule_excel_layout import weekly_hours_excel_formula

        formula = weekly_hours_excel_formula(0)
        self.assertTrue(formula.startswith("="))
        self.assertEqual(formula.count('="ΡΕΠΟ"'), 7)
        self.assertIn("TIME(INT(", formula)

    def test_interval_overnight_and_rest_excluded(self):
        from app.schedule_excel_layout import format_hours_minutes, weekly_declared_minutes

        rows = [
            {"hour_from": "22:00", "hour_to": "06:00", "shift_type": "ΕΡΓΑΣΙΑ"},
            {"hour_from": "09:00", "hour_to": "17:00", "shift_type": "ΑΝΑΠΑΥΣΗ/ΡΕΠΟ"},
            {"hour_from": "09:00", "hour_to": "17:00", "shift_type": "ADKAN"},
        ]
        self.assertEqual(format_hours_minutes(weekly_declared_minutes(rows)), "8:00")


if __name__ == "__main__":
    unittest.main()
