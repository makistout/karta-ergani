"""Δημιουργία Excel template εβδομαδιαίου ωραρίου (ένα φύλλο ανά ημέρα)."""

from __future__ import annotations

import sys
from datetime import date, timedelta

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from app.repo_entities import list_employees_for_employer
from app.repo_store import get_store_config

DAYS = [
    ("Δευτέρα", 0),
    ("Τρίτη", 1),
    ("Τετάρτη", 2),
    ("Πέμπτη", 3),
    ("Παρασκευή", 4),
    ("Σάββατο", 5),
    ("Κυριακή", 6),
]
HEADERS = ["ΑΦΜ", "Επώνυμο", "Όνομα", "Ενέργεια", "Από1", "Έως1", "Από2", "Έως2"]
TIME_COLS = [5, 6, 7, 8]


def build(store_id: int, out_path: str, week_monday: date | None = None) -> None:
    store = get_store_config(store_id)
    if not store:
        raise SystemExit(f"Δεν βρέθηκε κατάστημα id={store_id}")
    rows = list_employees_for_employer(store["employer_afm"], store["branch_aa"])

    today = date.today()
    monday = week_monday or (today + timedelta(days=(7 - today.weekday()) % 7 or 7))

    thin = Side(style="thin", color="D9E2F2")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    warn_fill = PatternFill("solid", fgColor="FFF2CC")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    wb = Workbook()
    info = wb.active
    info.title = "Οδηγίες"
    info["A1"] = f"Εβδομαδιαίο ωράριο - {store['name']}"
    info["A1"].font = Font(bold=True, size=14)
    sunday = monday + timedelta(days=6)
    info["A2"] = f"Εβδομάδα {monday.strftime('%d/%m/%Y')} - {sunday.strftime('%d/%m/%Y')}"
    info["A2"].font = Font(bold=True, size=11)
    instructions = [
        "",
        "Οδηγίες:",
        "1. Κάθε ημέρα είναι ξεχωριστό φύλλο (tab) κάτω-κάτω.",
        "2. Κάθε φύλλο δείχνει την πραγματική ημερομηνία στον τίτλο.",
        "3. Μία γραμμή ανά εργαζόμενο.",
        "4. Στήλη Ενέργεια: μόνο ΡΕΠΟ ή κενό.",
        "   - ΡΕΠΟ = ρεπό εκείνη την ημέρα (οι ώρες αγνοούνται).",
        "   - Κενό + ώρες συμπληρωμένες = αλλαγή ωραρίου.",
        "   - Κενό + κενές ώρες = καμία αλλαγή.",
        "5. Οι ώρες γράφονται ΩΩ:ΛΛ (π.χ. 09:00, 17:30).",
        "6. Για σπαστό ωράριο συμπλήρωσε και Από2/Έως2.",
    ]
    r = 3
    for text in instructions:
        info[f"A{r}"] = text
        r += 1
    info["A16"] = "Store ID"
    info["B16"] = store_id
    info["A17"] = "Employer AFM"
    info["B17"] = store["employer_afm"]
    info["A18"] = "Branch AA"
    info["B18"] = store["branch_aa"]
    info.column_dimensions["A"].width = 40
    info.column_dimensions["B"].width = 24

    for day_name, offset in DAYS:
        d = monday + timedelta(days=offset)
        ws = wb.create_sheet(f"{day_name} {d.strftime('%d-%m')}")
        ws["A1"] = f"{day_name} {d.strftime('%d/%m/%Y')}"
        ws["A1"].font = Font(bold=True, size=12)
        ws.merge_cells("A1:H1")
        ws["A2"] = "Ενέργεια: ΡΕΠΟ ή κενό  |  Ώρες: ΩΩ:ΛΛ"
        ws["A2"].font = Font(italic=True)
        ws["A2"].fill = warn_fill
        ws.merge_cells("A2:H2")

        for c_idx, title in enumerate(HEADERS, start=1):
            cell = ws.cell(row=3, column=c_idx, value=title)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border

        start_row = 4
        for i, emp in enumerate(rows, start=start_row):
            ws.cell(row=i, column=1, value=str(emp.get("afm") or ""))
            ws.cell(row=i, column=2, value=str(emp.get("eponymo") or ""))
            ws.cell(row=i, column=3, value=str(emp.get("onoma") or ""))
            for c in range(1, 9):
                cell = ws.cell(row=i, column=c)
                cell.border = border
                cell.alignment = center
            for tc in TIME_COLS:
                ws.cell(row=i, column=tc).number_format = "hh:mm"

        max_row = start_row + len(rows) - 1

        dv_action = DataValidation(
            type="list", formula1='"ΡΕΠΟ"', allow_blank=True, showDropDown=False
        )
        dv_action.errorTitle = "Μη έγκυρη τιμή"
        dv_action.error = "Επιτρέπεται μόνο ΡΕΠΟ ή κενό."
        dv_action.promptTitle = "Ενέργεια"
        dv_action.prompt = "Άφησε κενό ή επίλεξε ΡΕΠΟ"
        ws.add_data_validation(dv_action)
        dv_action.add(f"D{start_row}:D{max_row}")

        dv_time = DataValidation(type="time", allow_blank=True)
        dv_time.errorTitle = "Μη έγκυρη ώρα"
        dv_time.error = "Δώσε ώρα σε μορφή ΩΩ:ΛΛ (π.χ. 09:00)."
        dv_time.promptTitle = "Ώρα"
        dv_time.prompt = "Μορφή ΩΩ:ΛΛ, π.χ. 09:00"
        ws.add_data_validation(dv_time)
        for tc in TIME_COLS:
            col = get_column_letter(tc)
            dv_time.add(f"{col}{start_row}:{col}{max_row}")

        widths = {1: 14, 2: 22, 3: 18, 4: 12, 5: 10, 6: 10, 7: 10, 8: 10}
        for c, w in widths.items():
            ws.column_dimensions[get_column_letter(c)].width = w
        ws.freeze_panes = "A4"
        ws.auto_filter.ref = f"A3:H{max_row}"

    wb.save(out_path)
    print("OUT", out_path)
    print("WEEK", monday.isoformat())
    print("EMPLOYEES", len(rows))
    print("SHEETS", len(wb.sheetnames))


if __name__ == "__main__":
    sid = int(sys.argv[1]) if len(sys.argv) > 1 else 6
    out = (
        sys.argv[2]
        if len(sys.argv) > 2
        else r"c:\inetpub\wwwroot\erganios\scripts\weekly_schedule_template_stanotas_mykonos.xlsx"
    )
    build(sid, out)
